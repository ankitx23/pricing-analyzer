"""
Vexwire pricing pipeline - single-file version.

For each eBay seller account (bidallies, directauth, cellfeee):
  1. Scrape the account's last-month "Received as Seller" feedback (Item Title/ID/Price).
  2. Look up a comparable WatchCount price/title for each unique Item ID.
  3. Build a 2-sheet pricing analysis workbook (row-level Feedback + Avg Price by Item summary).

Usage:
    python run.py                              # runs all three accounts, one by one
    python run.py --account bidallies          # just one account
    python run.py --account cellfeee --skip-scrape   # reuse existing feedback.xlsx, redo pricing + analysis only

Requires Google Chrome installed. Missing Python packages are installed automatically on first run.
"""

import argparse
import importlib
import subprocess
import sys

REQUIRED_PACKAGES = {
    "selenium": "selenium",
    "bs4": "beautifulsoup4",
    "openpyxl": "openpyxl",
}


def _ensure_dependencies():
    missing = []
    for module_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(module_name)
        except ImportError:
            missing.append(pip_name)
    if missing:
        print(f"Installing missing dependencies: {', '.join(missing)}...", flush=True)
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])


_ensure_dependencies()

import json
import os
import re
import time
from collections import defaultdict
from datetime import date
from pathlib import Path
from urllib.parse import quote

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

REPO_ROOT = Path(__file__).resolve().parent
ACCOUNTS = ["bidallies", "directauth", "cellfeee"]
PROFILE_DIR = str(REPO_ROOT / ".chrome_profile")


# ---------------------------------------------------------------------------
# Chrome driver + WatchCount login
# ---------------------------------------------------------------------------

def _clear_stale_profile_locks():
    for name in ("SingletonLock", "SingletonSocket", "SingletonCookie", "DevToolsActivePort"):
        try:
            os.remove(os.path.join(PROFILE_DIR, name))
        except OSError:
            pass


def make_driver(retries=4, retry_delay=3):
    options = Options()
    options.add_argument(f"--user-data-dir={PROFILE_DIR}")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    last_error = None
    for attempt in range(retries + 1):
        _clear_stale_profile_locks()
        try:
            return webdriver.Chrome(options=options)
        except Exception as e:
            last_error = e
            print(f"    make_driver attempt {attempt + 1} failed: {e}")
            time.sleep(retry_delay)
    raise last_error


def is_logged_in(driver):
    return "CURRENT_USER = true" in driver.page_source or "CURRENT_USER = \"" in driver.page_source


def login(driver):
    driver.get("https://beta.watchcount.com/")
    time.sleep(2)
    if is_logged_in(driver):
        return

    email = os.environ.get("WATCHCOUNT_EMAIL")
    password = os.environ.get("WATCHCOUNT_PASSWORD")
    if not email or not password:
        print("WATCHCOUNT_EMAIL/PASSWORD not set - continuing without login (anonymous search still works).")
        return

    driver.get("https://beta.watchcount.com/login")
    time.sleep(3)
    driver.find_element(By.ID, "email").send_keys(email)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.XPATH, "//button[contains(@onclick, 'loginUser')]").click()
    time.sleep(4)


# ---------------------------------------------------------------------------
# WatchCount search/parsing + CAPTCHA handling
# ---------------------------------------------------------------------------

CAPTCHA_MARKER = 'id="challenge-section"'

WC_HIGHLIGHT_JS = """
const priceEl = document.querySelector('.find-results-new-item.item .price-container .price');
if (priceEl) {
    priceEl.scrollIntoView({block: 'center'});
    priceEl.style.outline = '3px solid red';
    priceEl.style.borderRadius = '4px';
    priceEl.style.backgroundColor = 'rgba(255,0,0,0.08)';
}
"""


def wc_build_search_url(query, offset=0):
    keywords = str(query).replace(" ", "+")
    url = f"https://www.watchcount.com/live/{keywords}/-/all"
    if offset:
        url += f"?offset={offset}"
    return url


def wc_build_sold_url(item_id):
    return f"https://www.watchcount.com/sold/{item_id}/-/all"


def wc_parse_cards(html):
    soup = BeautifulSoup(html, "html.parser")
    cards = soup.select(".find-results-new-item.item")
    items = []
    for card in cards:
        text_block = card.select_one(".col.text-center.bold-text")
        watchers = sold = avg = None
        if text_block:
            text = text_block.get_text(" ", strip=True)
            wm = re.search(r"Watchers:\s*([\d,]+)", text)
            sm = re.search(r"Sold:\s*([\d,]+)", text)
            am = re.search(r"Average:\s*([^*]+)", text)
            watchers = wm.group(1) if wm else "N/A"
            sold = sm.group(1) if sm else "N/A"
            avg = re.sub(r"[^\x00-\x7F]+", "", am.group(1)).strip() if am else "N/A"

        price_container = card.select_one(".price-container")
        price = "N/A"
        free_shipping = False
        if price_container:
            price_el = price_container.select_one(".price")
            price = price_el.get_text(strip=True) if price_el else "N/A"
            free_shipping = "Free Shipping" in price_container.get_text()

        general = card.select_one(".general-info-container")
        title = condition = "N/A"
        item_number = "N/A"
        if general:
            direct_rows = general.find_all("div", class_="row", recursive=False)
            row_texts = [r.get_text(" ", strip=True) for r in direct_rows]
            if len(row_texts) > 0:
                title = row_texts[0]
            if len(row_texts) > 1:
                condition = row_texts[1]
            for span in general.select("span"):
                t = span.get_text(strip=True)
                if t.isdigit() and len(t) >= 9:
                    item_number = t
                    break

        items.append({
            "watchCount": watchers,
            "quantitySold": sold,
            "oneUnitEvery": avg,
            "priceFormatted": price,
            "freeShipping": free_shipping,
            "condition": condition,
            "id": item_number,
            "title": title,
        })
    return items


def wc_is_captcha_page(driver):
    return CAPTCHA_MARKER in driver.page_source


def wc_wait_for_captcha_solved(driver, poll_seconds=3, reminder_every=30, max_wait=60):
    print("    CAPTCHA challenge detected - please solve it in the browser window. Waiting (not navigating away)...", flush=True)
    waited = 0
    while wc_is_captcha_page(driver):
        if waited >= max_wait:
            print(f"    Gave up waiting for CAPTCHA after {max_wait}s, continuing anyway.", flush=True)
            return
        time.sleep(poll_seconds)
        waited += poll_seconds
        if waited % reminder_every == 0:
            print(f"    still waiting on CAPTCHA ({waited}s)...", flush=True)
    print("    CAPTCHA cleared, continuing.", flush=True)


def wc_find_matching_item(driver, url, item_id):
    driver.get(url)
    time.sleep(4)
    if wc_is_captcha_page(driver):
        wc_wait_for_captcha_solved(driver)
        time.sleep(1)
    items = wc_parse_cards(driver.page_source)
    for item in items:
        if str(item.get("id")) == str(item_id):
            try:
                driver.execute_script(WC_HIGHLIGHT_JS)
            except Exception:
                pass
            time.sleep(1)
            return {"price": item.get("priceFormatted", "N/A"), "title": item.get("title", "N/A")}
    return None


def wc_lookup_price(driver, item_id):
    result = wc_find_matching_item(driver, wc_build_search_url(str(item_id)), item_id)
    if result is not None:
        return result
    result = wc_find_matching_item(driver, wc_build_sold_url(item_id), item_id)
    if result is not None:
        return result
    return {"price": "No match", "title": "No match"}


# ---------------------------------------------------------------------------
# Step 1: scrape eBay feedback
# ---------------------------------------------------------------------------

PAGE_LIMIT = 25
FEEDBACK_FILTER = "feedback_page:RECEIVED_AS_SELLER,period:ONE_MONTH,include_automated_feedback:true"


def fb_build_url(username, page_id):
    base = f"https://www.ebay.com/fdbk/feedback_profile/{username}"
    return f"{base}?username={username}&sort=TIME&filter={quote(FEEDBACK_FILTER)}&page_id={page_id}&limit={PAGE_LIMIT}"


def fb_parse_page(html):
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    for tr in soup.select("tr[data-feedback-id]"):
        feedback_id = tr["data-feedback-id"]
        item_div = tr.select_one(".card__item")
        price_el = tr.select_one(".card__price span")
        if not item_div or not price_el:
            continue
        link = item_div.select_one("a[href*='/itm/']")
        if not link:
            continue
        item_id = link.get_text(strip=True)
        spans = item_div.find_all("span", recursive=False)
        title = spans[0].get_text(strip=True) if spans else ""
        title = re.sub(r"\s*\(#\s*$", "", title).strip()
        price = price_el.get_text(strip=True)
        rows.append((feedback_id, title, item_id, price))
    return rows


def fb_get_total_count(html):
    m = re.search(r"([\d,]+)\s*Feedback received", html)
    return int(m.group(1).replace(",", "")) if m else None


def fb_format_workbook(ws):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in range(2, ws.max_row + 1):
        item_id_cell = ws.cell(row=row, column=2)
        item_id_cell.value = str(item_id_cell.value)
        item_id_cell.number_format = "@"
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if row % 2 == 0:
                cell.fill = stripe_fill
            if col == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 3:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    max_title_len = max((len(str(ws.cell(row=r, column=1).value or "")) for r in range(2, ws.max_row + 1)), default=10)
    ws.column_dimensions["A"].width = min(max_title_len + 4, 90)
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.row_dimensions[1].height = 20

    red = Side(style="thick", color="FF0000")
    last_row = ws.max_row
    last_col = ws.max_column
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            existing = cell.border
            top = red if row == 1 else existing.top
            bottom = red if row == last_row else existing.bottom
            left = red if col == 1 else existing.left
            right = red if col == last_col else existing.right
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def fb_load_existing(output_path):
    try:
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        rows = []
        for r in range(2, ws.max_row + 1):
            title = ws.cell(row=r, column=1).value
            item_id = ws.cell(row=r, column=2).value
            price = ws.cell(row=r, column=3).value
            if title is None:
                continue
            rows.append((title, item_id, price))
        return rows
    except FileNotFoundError:
        return []


def fb_load_seen_ids(seen_ids_path):
    try:
        with open(seen_ids_path) as f:
            return {line.strip() for line in f if line.strip()}
    except FileNotFoundError:
        return set()


def fb_append_seen_id(seen_ids_path, feedback_id):
    with open(seen_ids_path, "a") as f:
        f.write(f"{feedback_id}\n")


def fb_load_checkpoint(checkpoint_path):
    try:
        with open(checkpoint_path) as f:
            return int(f.read().strip())
    except FileNotFoundError:
        return 0


def fb_save_checkpoint(checkpoint_path, page_id):
    with open(checkpoint_path, "w") as f:
        f.write(str(page_id))


def scrape_feedback(username, output_dir, max_pages_this_run):
    output_path = output_dir / "feedback.xlsx"
    checkpoint_path = output_dir / "checkpoint.txt"
    seen_ids_path = output_dir / "seen_feedback_ids.txt"

    all_rows = fb_load_existing(output_path)
    seen_ids = fb_load_seen_ids(seen_ids_path)
    start_page = fb_load_checkpoint(checkpoint_path) + 1
    print(f"Resuming from page {start_page}, already have {len(all_rows)} rows ({len(seen_ids)} feedback IDs seen)", flush=True)

    driver = make_driver()
    pages_done_this_run = 0
    page_id = start_page
    try:
        total = None
        empty_streak = 0
        while pages_done_this_run < max_pages_this_run:
            url = fb_build_url(username, page_id)
            print(f"Fetching page {page_id}: {url}", flush=True)
            driver.get(url)
            time.sleep(6)
            html = driver.page_source

            if total is None:
                total = fb_get_total_count(html)
                print(f"Total feedback: {total}", flush=True)

            rows = fb_parse_page(html)
            new_rows = [r for r in rows if r[0] not in seen_ids]
            for r in new_rows:
                seen_ids.add(r[0])
                fb_append_seen_id(seen_ids_path, r[0])
            print(f"  -> {len(rows)} rows on page, {len(new_rows)} new, total so far {len(all_rows) + len(new_rows)}", flush=True)
            all_rows.extend((title, item_id, price) for _fid, title, item_id, price in new_rows)
            fb_save_checkpoint(checkpoint_path, page_id)
            pages_done_this_run += 1

            if total is not None and len(all_rows) >= total:
                print("Reached total, done.", flush=True)
                break
            if not rows:
                empty_streak += 1
                if empty_streak >= 2:
                    print("Two empty pages in a row, stopping (done).", flush=True)
                    break
            else:
                empty_streak = 0
            page_id += 1
    finally:
        driver.quit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback"
    ws.append(["Item Title", "Item ID", "Price"])
    for title, item_id, price in all_rows:
        ws.append([title, item_id, price])
    fb_format_workbook(ws)
    wb.save(output_path)
    print(f"\nSaved {len(all_rows)} rows to {output_path} (checkpoint at page {page_id})", flush=True)
    return output_path


# ---------------------------------------------------------------------------
# Step 2: look up WatchCount prices
# ---------------------------------------------------------------------------

WAIT_SECONDS = 5
PRICE_HEADER = "WatchCount Price"
TITLE_HEADER = "WatchCount Title"


def wcp_load_cache(cache_path):
    try:
        with open(cache_path) as f:
            cache = json.load(f)
    except FileNotFoundError:
        return {}
    for item_id, value in cache.items():
        if isinstance(value, str):
            cache[item_id] = {"price": value, "title": "N/A"}
    return cache


def wcp_save_cache(cache_path, cache):
    with open(cache_path, "w") as f:
        json.dump(cache, f, indent=2)


def wcp_find_or_create_column(ws, header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=header)
    return col


def fetch_watchcount_prices(feedback_path, cache_path, max_new=10_000):
    wb = openpyxl.load_workbook(feedback_path)
    ws = wb.active
    price_col = wcp_find_or_create_column(ws, PRICE_HEADER)
    title_col = wcp_find_or_create_column(ws, TITLE_HEADER)

    item_id_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Item ID":
            item_id_col = col
            break
    if item_id_col is None:
        raise RuntimeError("Could not find 'Item ID' column")

    unique_ids = []
    seen = set()
    for row in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row, column=item_id_col).value
        if item_id is None or str(item_id) in seen:
            continue
        seen.add(str(item_id))
        unique_ids.append(str(item_id))

    cache = wcp_load_cache(cache_path)
    print(f"{len(unique_ids)} unique item IDs, {len(cache)} already cached", flush=True)

    driver = make_driver()
    new_lookups = 0
    try:
        for i, item_id in enumerate(unique_ids):
            if item_id in cache:
                continue
            if new_lookups >= max_new:
                print(f"Reached limit of {max_new} new lookups this run. Stopping.", flush=True)
                break

            print(f"[{i + 1}/{len(unique_ids)}] Looking up {item_id}...", flush=True)
            try:
                result = wc_lookup_price(driver, item_id)
            except Exception as e:
                print(f"    ERROR: {e}", flush=True)
                result = {"price": "N/A", "title": "N/A"}

            cache[item_id] = result
            new_lookups += 1
            print(f"    -> {result['price']} | {result['title']}", flush=True)
            wcp_save_cache(cache_path, cache)

            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=item_id_col).value) == item_id:
                    ws.cell(row=row, column=price_col, value=result["price"])
                    ws.cell(row=row, column=title_col, value=result["title"])
            wb.save(feedback_path)

            time.sleep(WAIT_SECONDS)
    finally:
        driver.quit()

    for row in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row, column=item_id_col).value
        if item_id is not None and str(item_id) in cache:
            result = cache[str(item_id)]
            ws.cell(row=row, column=price_col, value=result["price"])
            ws.cell(row=row, column=title_col, value=result["title"])
    wb.save(feedback_path)
    print(f"Done. {len(cache)}/{len(unique_ids)} unique item IDs resolved. Saved to {feedback_path}", flush=True)


# ---------------------------------------------------------------------------
# Step 3: build the analysis workbook
# ---------------------------------------------------------------------------

ANALYSIS_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
ANALYSIS_HEADER_FONT = Font(bold=True, color="FFFFFF")

FEEDBACK_HEADERS = [
    "Item Title", "Item ID", "Price", "WatchCount Price", "WatchCount Title",
    "Watch * 1.5", "Price/WatchCount Price", "Price/WatchCount Price (Rounded)", "Avg Per Unit",
]
SUMMARY_HEADERS = ["Item title", "Item ID", "Qty", "Average Price", "WatchCount Price", "WatchCount Title"]


def an_parse_money(value):
    if value is None:
        return None
    match = re.search(r"[\d,]+\.?\d*", str(value))
    if not match:
        return None
    return float(match.group(0).replace(",", ""))


def an_style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = ANALYSIS_HEADER_FILL
        cell.font = ANALYSIS_HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def an_autosize(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def build_analysis(source_path, output_path):
    src_wb = openpyxl.load_workbook(source_path)
    src_ws = src_wb.active
    src_headers = [c.value for c in src_ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(src_headers)}

    required = ["Item Title", "Item ID", "Price", "WatchCount Price", "WatchCount Title"]
    missing = [h for h in required if h not in col]
    if missing:
        raise RuntimeError(f"{source_path} is missing columns {missing} - the WatchCount price lookup step must run first.")

    out_wb = openpyxl.Workbook()
    feedback_ws = out_wb.active
    feedback_ws.title = "Feedback"
    feedback_ws.append(FEEDBACK_HEADERS)

    item_rows = defaultdict(list)  # item_id -> list of (title, price_float, watchcount_price_str, watchcount_title)

    out_row = 2
    for r in range(2, src_ws.max_row + 1):
        title = src_ws.cell(row=r, column=col["Item Title"]).value
        item_id = str(src_ws.cell(row=r, column=col["Item ID"]).value)
        price_raw = src_ws.cell(row=r, column=col["Price"]).value
        wc_price_raw = src_ws.cell(row=r, column=col["WatchCount Price"]).value
        wc_title = src_ws.cell(row=r, column=col["WatchCount Title"]).value

        price = an_parse_money(price_raw)
        wc_price = an_parse_money(wc_price_raw)

        feedback_ws.cell(row=out_row, column=1, value=title)
        id_cell = feedback_ws.cell(row=out_row, column=2, value=item_id)
        id_cell.number_format = "@"
        feedback_ws.cell(row=out_row, column=3, value=price)
        feedback_ws.cell(row=out_row, column=4, value=wc_price)
        feedback_ws.cell(row=out_row, column=5, value=wc_title)
        feedback_ws.cell(row=out_row, column=6, value=f"=D{out_row}*1.5")
        feedback_ws.cell(row=out_row, column=7, value=f"=C{out_row}/D{out_row}")
        feedback_ws.cell(row=out_row, column=8, value=f"=ROUND(G{out_row},0)")
        feedback_ws.cell(row=out_row, column=9, value=f"=C{out_row}/H{out_row}")

        item_rows[item_id].append((title, price, wc_price_raw, wc_title))
        out_row += 1

    an_style_header(feedback_ws, len(FEEDBACK_HEADERS))
    an_autosize(feedback_ws, {
        "A": 81, "B": 14, "C": 10, "D": 16, "E": 77,
        "F": 13, "G": 22, "H": 22, "I": 14,
    })

    summary_ws = out_wb.create_sheet("Avg Price by Item")
    summary_ws.append(SUMMARY_HEADERS)

    row = 2
    for item_id in sorted(item_rows.keys(), key=int):
        entries = item_rows[item_id]
        title = entries[0][0]
        qty = len(entries)
        avg_price = round(sum(e[1] for e in entries) / qty, 2)
        wc_price_display = entries[0][2]
        wc_title = entries[0][3]

        summary_ws.cell(row=row, column=1, value=title)
        id_cell = summary_ws.cell(row=row, column=2, value=item_id)
        id_cell.number_format = "@"
        summary_ws.cell(row=row, column=3, value=qty)
        summary_ws.cell(row=row, column=4, value=avg_price)
        summary_ws.cell(row=row, column=5, value=wc_price_display)
        summary_ws.cell(row=row, column=6, value=wc_title)
        row += 1

    an_style_header(summary_ws, len(SUMMARY_HEADERS))
    an_autosize(summary_ws, {"A": 81, "B": 14, "C": 6, "D": 14, "E": 16, "F": 77})

    out_wb.save(output_path)
    print(f"Feedback rows: {out_row - 2}")
    print(f"Unique items: {row - 2}")
    print(f"Saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def run_account(username, skip_scrape, max_pages):
    output_dir = REPO_ROOT / "output" / username
    output_dir.mkdir(parents=True, exist_ok=True)
    feedback_path = output_dir / "feedback.xlsx"

    print(f"\n########## {username} ##########", flush=True)

    if skip_scrape and feedback_path.exists():
        print(f"[{username}] Step 1/3: skipping scrape, reusing existing feedback.xlsx", flush=True)
    else:
        print(f"[{username}] Step 1/3: scraping eBay feedback...", flush=True)
        scrape_feedback(username, output_dir, max_pages)

    print(f"[{username}] Step 2/3: looking up WatchCount prices...", flush=True)
    print("    Chrome will open. If WatchCount shows a verification challenge, solve it in that window - "
          "the script detects it and continues automatically (waits up to 60s).", flush=True)
    cache_path = output_dir / "watchcount_prices.json"
    fetch_watchcount_prices(feedback_path, cache_path)

    print(f"[{username}] Step 3/3: building analysis workbook...", flush=True)
    today = date.today().isoformat()
    analysis_path = output_dir / f"{username}_analysis_{today}.xlsx"
    build_analysis(feedback_path, analysis_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--account", default="all", help=f"One of {ACCOUNTS}, or 'all' (default)")
    parser.add_argument("--skip-scrape", action="store_true", help="Reuse existing feedback.xlsx instead of scraping live")
    parser.add_argument("--max-pages", type=int, default=55, help="Max feedback pages to fetch per run (resumable via checkpoint)")
    args = parser.parse_args()

    targets = ACCOUNTS if args.account == "all" else [args.account]
    for username in targets:
        if username not in ACCOUNTS:
            raise SystemExit(f"Unknown account {username!r}. Options: {ACCOUNTS}")

    print(
        "NOTE: Chrome will open visibly during each account's WatchCount price-lookup step. If WatchCount "
        "shows a verification challenge, solve it in that window - the script waits for you (up to 60s) and "
        "continues automatically. This step cannot run fully unattended.\n",
        flush=True,
    )

    failed = []
    for username in targets:
        try:
            run_account(username, args.skip_scrape, args.max_pages)
        except Exception as e:
            print(f"\n[{username}] FAILED: {e}", flush=True)
            print(f"[{username}] Skipping to the next account. Re-run with --account {username} "
                  f"--skip-scrape once this is fixed to pick up where it left off.", flush=True)
            failed.append(username)

    if failed:
        print(f"\nDone, but {len(failed)}/{len(targets)} account(s) failed: {', '.join(failed)}. "
              f"See errors above. Others' output is still under output/<account>/.", flush=True)
        raise SystemExit(1)

    print("\nAll done. Analysis workbooks are under output/<account>/.", flush=True)


if __name__ == "__main__":
    main()
